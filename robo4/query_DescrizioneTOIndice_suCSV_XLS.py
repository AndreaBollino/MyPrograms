import cx_Oracle
#import os
import csv
from openpyxl import Workbook

#un = os.environ.get('USER_NAME')
#pw = os.environ.get('PASSWORD')

cx_Oracle.init_oracle_client(config_dir="C:\\oracle\\ora12R2CL_32\\network\\admin")
connection = cx_Oracle.connect(dsn="OraWV_NPWAADVISORYAPP_PROD.world", encoding="UTF-8")
cursor = connection.cursor()
query = """
WITH totale AS
  (SELECT r4s.DESCRIPTION AS R4_SESSION_DESCRIPTION,
    s.INDICE_OUTPUT       AS BS_SESSION_SCHED_INDICE_OUTPUT,
    r4t.id                AS R4_TARGET_ID,
    bsw.START_DATE        AS data_partenza
  FROM NPWACDR4.BS_SESSION_SCHED s
  INNER JOIN Npwacdr4.Bs_Session_Progress P
  ON P.Id_Session_Sched = S.Id
  LEFT JOIN NPWACDR4.R4_TARGET r4t
  ON r4t.id = s.CODICE_SESSIONE
  LEFT JOIN NPWACDR4.R4_SESSION r4s
  ON r4s.id = r4t.SESSION_ID
  JOIN NPWACDR4.BS_SESSION_WORKER_PROGRESS bsw
  ON p.id              = bsw.ID_SESSION_PROGRESS
  WHERE 1              =1
  AND r4s.DESCRIPTION IS NOT NULL
  AND  (LOWER (R4s.Description) = LOWER ('CA Sottopeso US Equity SETT 2023 bis') or
        LOWER (R4s.Description) = LOWER ('CA Sottopeso US Equity SETT 2023') or
        LOWER (R4s.Description) = LOWER ('CA Top di Gamma Aderenza < 7 SETT 2023') or
        LOWER (R4s.Description) = LOWER ('CA Top di Gamma Adeguatezza Rischio MKT SETT 2023') or
        LOWER (R4s.Description) = LOWER ('CB Top di Gamma Adeguatezza Rischio MKT SETT 2023')
)
  --LIKE 'CA Top di Gamma Aderenza < 7 LUG-AGO 23'
  --AND R4s.Description LIKE :description
    --and s.INDICE_OUTPUT  in ('bsrobo4-risp-picking-2507','bsrobo4-risp-picking-2506')
    --order by r4t.id desc-- as R4_TARGET_ID
  ORDER BY bsw.START_DATE DESC
  )
SELECT R4_SESSION_DESCRIPTION AS DESCRIZIONE,
  --BS_SESSION_SCHED_INDICE_OUTPUT,
  R4_TARGET_ID AS INDICE
FROM totale
GROUP BY R4_SESSION_DESCRIPTION,
 -- BS_SESSION_SCHED_INDICE_OUTPUT,
  R4_TARGET_ID
"""
cursor.execute(query)

# Recupero dei risultati
results = cursor.fetchall()

# Estrazione dei nomi delle colonne
column_names = [col[0] for col in cursor.description]

# Lettura dei risultati
for row in results:
    print(row)

output_file = "indice.csv"

# Scrittura dei risultati su un file CSV
with open(output_file, "w", newline="") as csvfile:
    writer = csv.writer(csvfile, delimiter=";")
    
    # Scrittura dei nomi delle colonne come prima riga
    writer.writerow(column_names)

    # Scrittura dei dati
    for row in results:
        writer.writerow(row)

# Creazione del foglio di lavoro Excel
workbook = Workbook()
sheet = workbook.active

# Scrittura dei nomi delle colonne nel foglio di lavoro
for col_idx, column_name in enumerate(column_names, start=1):
    sheet.cell(row=1, column=col_idx, value=column_name)

# Scrittura dei risultati della query nel foglio di lavoro
row_idx = 2
for row in results:
    for col_idx, value in enumerate(row, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=value)
    row_idx += 1

# Salvataggio del foglio di lavoro
workbook.save('Indice.xlsx')

# Chiusura della connessione
cursor.close()
connection.close()

print("Query eseguita con successo. I risultati sono stati scritti su", output_file)