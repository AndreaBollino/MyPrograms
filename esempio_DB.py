import cx_Oracle
#import os
import csv
from openpyxl import Workbook

#un = os.environ.get('USER_NAME')
#pw = os.environ.get('PASSWORD')

cx_Oracle.init_oracle_client(config_dir="C:\\oracle\\ora12R2CL_32\\network\\admin")
connection = cx_Oracle.connect(dsn="OraWV_NPWAADVISORYAPP_PROD.world", encoding="UTF-8")
cursor = connection.cursor()
query = "select * from NPWAPFP.tbl_guida_staging order by TS_AGGIORNAMENTO_VIEW desc "
cursor.execute(query)

# Recupero dei risultati
results = cursor.fetchall()

# Estrazione dei nomi delle colonne
column_names = [col[0] for col in cursor.description]

# Lettura dei risultati
for row in results:
    print(row)

output_file = "query.csv"

# Scrittura dei risultati su un file CSV
with open(output_file, "w", newline="") as csvfile:
    writer = csv.writer(csvfile, delimiter=";")
    
    # Scrittura dei nomi delle colonne come prima riga
    writer.writerow(column_names)


    # Scrittura dei dati
    for row in results:
        writer.writerow(row)

# Creazione del foglio di lavoro Excel
workbook = Workbook()
sheet = workbook.active

# Scrittura dei nomi delle colonne nel foglio di lavoro
for col_idx, column_name in enumerate(column_names, start=1):
    sheet.cell(row=1, column=col_idx, value=column_name)

# Scrittura dei risultati della query nel foglio di lavoro
row_idx = 2
for row in results:
    for col_idx, value in enumerate(row, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=value)
    row_idx += 1

# Salvataggio del foglio di lavoro
workbook.save('query2.xlsx')

# Chiusura della connessione
cursor.close()
connection.close()

print("Query eseguita con successo. I risultati sono stati scritti su", output_file)
